from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .config import RAW_COLUMNS, RAW_FALLBACK_SHEETS, RAW_SOURCE_SHEET


def _make_headers_unique(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    unique_headers: list[str] = []
    for index, header in enumerate(headers):
        base = header if header is not None else f"unnamed_{index}"
        count = seen.get(base, 0)
        if count:
            unique_headers.append(f"{base}_{count}")
        else:
            unique_headers.append(base)
        seen[base] = count + 1
    return unique_headers


def _load_sheet_as_dataframe(workbook, sheet_name: str) -> pd.DataFrame:
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    header = _make_headers_unique(list(next(rows)))
    return pd.DataFrame(rows, columns=header)


def _sheet_score(dataframe: pd.DataFrame) -> int:
    preferred_columns = [
        RAW_COLUMNS["date_exam"],
        RAW_COLUMNS["physician"],
        RAW_COLUMNS["traction"],
        RAW_COLUMNS["atract"],
        RAW_COLUMNS["major_diameter"],
    ]
    return sum(column in dataframe.columns for column in preferred_columns)


def load_raw_workbook(path: str | Path) -> pd.DataFrame:
    workbook = load_workbook(filename=Path(path), read_only=True, data_only=True)
    candidate_sheets = []
    for sheet_name in [RAW_SOURCE_SHEET, *RAW_FALLBACK_SHEETS, *workbook.sheetnames]:
        if sheet_name in workbook.sheetnames and sheet_name not in candidate_sheets:
            candidate_sheets.append(sheet_name)

    best_dataframe: pd.DataFrame | None = None
    best_score = -1
    for sheet_name in candidate_sheets:
        dataframe = _load_sheet_as_dataframe(workbook, sheet_name)
        score = _sheet_score(dataframe)
        if score > best_score:
            best_dataframe = dataframe
            best_score = score

    if best_dataframe is None:
        raise ValueError(f"No readable worksheet found in {path}")
    return best_dataframe


def normalize_missing_values(dataframe: pd.DataFrame) -> pd.DataFrame:
    cleaned = dataframe.copy()
    cleaned = cleaned.replace({"NULL": pd.NA, "#VALUE!": pd.NA, "": pd.NA})
    return cleaned
